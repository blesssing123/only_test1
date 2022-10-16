# name = "alex"
# # print("hello", name)
# # def sayhi(n):
# #     print("hi", n)
# import re
# print(re.search("1\d{9}","15926876212"))
# import time
# start_time = time.strftime("%Y-%m-%d %H:%M:%S")
# print(start_time)
# print(time.strptime(start_time, '%Y-%m-%d %H:%M:%S'))
import datetime
from openpyxl import Workbook, load_workbook

# wb = Workbook()
# sheet = wb.active
# sheet.title = '测试表1'
# sheet['E3'] = '张三'
# sheet['E4'] = '李四'
# sheet['a2'] = datetime.datetime.now().strftime('%Y-%m-%d')
# sheet.append(['王五', 170 , '70kg'])
# print(sheet.title)
# wb.save('excel_test.xlsx')
wb = load_workbook('excel_test.xlsx')
print(wb._sheets)
sheet = wb['测试表1']
print(sheet['E3'].value)
